"""Unit tests for services/shared/xlsx_compose.py

All tests are self-contained:
- each test writes to a temp file (via tmp_path fixture)
- tests load the written workbook with openpyxl to assert structural truth
- no mocking of filesystem or openpyxl — we want to catch serialisation bugs

Run: python -m pytest tests/unit/test_xlsx_compose.py -v
"""
from __future__ import annotations

import openpyxl
import pytest

from services.shared.xlsx_compose import compose_xlsx, validate_xlsx_spec


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _load(path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(str(path))


# ---------------------------------------------------------------------------
# 1. Simple one-sheet with headers and rows
# ---------------------------------------------------------------------------


def test_simple_one_sheet_with_headers_and_rows(tmp_path):
    out = tmp_path / "simple.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Summary",
                "headers": ["Metric", "Value", "Notes"],
                "rows": [
                    ["Revenue", 1_000_000, "Q1 actual"],
                    ["Costs", 750_000, "Q1 actual"],
                    ["Profit", 250_000, "Q1 derived"],
                ],
            }
        ]
    }
    result = compose_xlsx(spec, out)

    assert result["sheets"] == 1
    assert result["cells_written"] > 0
    assert out.exists()

    wb = _load(out)
    ws = wb["Summary"]

    # Headers in row 1
    assert ws["A1"].value == "Metric"
    assert ws["B1"].value == "Value"
    assert ws["C1"].value == "Notes"

    # First data row
    assert ws["A2"].value == "Revenue"
    assert ws["B2"].value == 1_000_000
    assert ws["C2"].value == "Q1 actual"

    # Third data row
    assert ws["A4"].value == "Profit"


# ---------------------------------------------------------------------------
# 2. Formulas written as strings starting with "="
# ---------------------------------------------------------------------------


def test_formulas_written_as_strings(tmp_path):
    out = tmp_path / "formulas.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Model",
                "headers": ["Item", "Qty", "Price", "Total"],
                "rows": [
                    ["Widget A", 10, 25.0, None],
                    ["Widget B", 5, 40.0, None],
                    ["", "", "Sum", None],
                ],
                "formulas": {
                    "D2": "=B2*C2",
                    "D3": "=B3*C3",
                    "D4": "=SUM(D2:D3)",
                },
            }
        ]
    }
    result = compose_xlsx(spec, out)
    assert result["cells_written"] > 0

    wb = _load(out)
    ws = wb["Model"]

    # openpyxl stores formulas as-is when load_workbook(data_only=False)
    assert ws["D2"].value == "=B2*C2"
    assert ws["D3"].value == "=B3*C3"
    assert ws["D4"].value == "=SUM(D2:D3)"

    # All three must start with "="
    for cell_ref in ("D2", "D3", "D4"):
        val = ws[cell_ref].value
        assert isinstance(val, str) and val.startswith("="), (
            f"Expected formula string at {cell_ref}, got {val!r}"
        )


# ---------------------------------------------------------------------------
# 3. Column widths applied
# ---------------------------------------------------------------------------


def test_column_widths_applied(tmp_path):
    out = tmp_path / "widths.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Sheet1",
                "headers": ["Long Column Name", "Short"],
                "rows": [["alpha", "beta"]],
                "column_widths": {"A": 30.0, "B": 12.5},
            }
        ]
    }
    compose_xlsx(spec, out)

    wb = _load(out)
    ws = wb["Sheet1"]

    assert ws.column_dimensions["A"].width == pytest.approx(30.0)
    assert ws.column_dimensions["B"].width == pytest.approx(12.5)


# ---------------------------------------------------------------------------
# 4. Freeze panes set
# ---------------------------------------------------------------------------


def test_freeze_panes_set(tmp_path):
    out = tmp_path / "freeze.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Data",
                "headers": ["Date", "Amount"],
                "rows": [["2026-01-01", 100]],
                "freeze_panes": "A2",
            }
        ]
    }
    compose_xlsx(spec, out)

    wb = _load(out)
    ws = wb["Data"]
    assert ws.freeze_panes == "A2"


# ---------------------------------------------------------------------------
# 5. Brooker style — header row has navy fill
# ---------------------------------------------------------------------------


def test_brooker_style_header_row_navy(tmp_path):
    out = tmp_path / "style.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Styled",
                "headers": ["Col A", "Col B", "Col C"],
                "rows": [["x", "y", "z"]],
                "style": "brooker",
            }
        ]
    }
    compose_xlsx(spec, out)

    wb = _load(out)
    ws = wb["Styled"]

    # Header is row 1 (no metadata title on this spec)
    header_cell = ws["A1"]
    fill = header_cell.fill
    # PatternFill fgColor.rgb includes alpha prefix "FF" + hex
    assert fill.fgColor.rgb.upper() in ("FF0F3D5C", "0F3D5C"), (
        f"Expected navy header fill, got {fill.fgColor.rgb!r}"
    )
    assert header_cell.font.bold is True
    # White text
    assert header_cell.font.color.rgb.upper() in ("FFFFFFFF", "FFFFFF"), (
        f"Expected white font, got {header_cell.font.color.rgb!r}"
    )


# ---------------------------------------------------------------------------
# 6. Validation rejects oversized sheet name (sanitised to 31 chars)
# ---------------------------------------------------------------------------


def test_validate_rejects_oversized_sheet_name():
    long_name = "A" * 35  # 35 chars — over the 31-char Excel limit
    spec = {
        "sheets": [
            {
                "name": long_name,
                "rows": [["a", "b"]],
            }
        ]
    }
    normalised = validate_xlsx_spec(spec)
    assert len(normalised["sheets"][0]["name"]) == 31
    assert normalised["sheets"][0]["name"] == "A" * 31


# ---------------------------------------------------------------------------
# 7. Validation rejects row with more cells than headers
# ---------------------------------------------------------------------------


def test_validate_rejects_row_length_mismatch_with_headers():
    spec = {
        "sheets": [
            {
                "name": "Bad",
                "headers": ["A", "B", "C"],   # 3 headers
                "rows": [
                    ["ok", "ok", "ok"],         # fine
                    ["too", "many", "cols", "extra"],  # 4 cells — bad
                ],
            }
        ]
    }
    with pytest.raises(ValueError, match="headers has 3 columns"):
        validate_xlsx_spec(spec)


# ---------------------------------------------------------------------------
# 8. Chart embedded in named sheet
# ---------------------------------------------------------------------------


def test_chart_embedded_in_named_sheet(tmp_path):
    out = tmp_path / "chart.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Revenue",
                "headers": ["Month", "Amount"],
                "rows": [
                    ["Jan", 100_000],
                    ["Feb", 120_000],
                    ["Mar", 115_000],
                    ["Apr", 130_000],
                ],
            }
        ],
        "charts": [
            {
                "sheet": "Revenue",
                "kind": "bar",
                "data_range": "Revenue!B1:B5",
                "category_range": "Revenue!A2:A5",
                "title": "Monthly Revenue",
                "anchor": "D2",
            }
        ],
    }
    result = compose_xlsx(spec, out)
    assert result["charts"] == 1

    wb = _load(out)
    ws = wb["Revenue"]
    assert len(ws._charts) == 1, f"Expected 1 chart on sheet, found {len(ws._charts)}"


# ---------------------------------------------------------------------------
# 9. Conditional data_bar rule applied to range
# ---------------------------------------------------------------------------


def test_conditional_data_bar_applied(tmp_path):
    out = tmp_path / "cond.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Metrics",
                "headers": ["KPI", "Value"],
                "rows": [
                    ["Revenue growth", 0.12],
                    ["Cost reduction", 0.08],
                    ["Headcount efficiency", 0.95],
                ],
                "conditional_format": [
                    {"range": "B2:B4", "type": "data_bar", "color": "blue"}
                ],
            }
        ]
    }
    compose_xlsx(spec, out)

    wb = _load(out)
    ws = wb["Metrics"]

    # openpyxl stores conditional formatting keyed by range string
    cf_ranges = list(ws.conditional_formatting)
    assert len(cf_ranges) > 0, "No conditional formatting rules found"

    found = False
    for cf_range in cf_ranges:
        if "B2" in str(cf_range) and "B4" in str(cf_range):
            found = True
            break
    assert found, f"Expected a rule on B2:B4, found ranges: {cf_ranges}"


# ---------------------------------------------------------------------------
# 10. Metadata applied to workbook properties
# ---------------------------------------------------------------------------


def test_metadata_applied(tmp_path):
    out = tmp_path / "meta.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Report",
                "headers": ["Item"],
                "rows": [["row1"]],
            }
        ],
        "metadata": {
            "title": "CAC Monthly Report",
            "author": "CAC Agent",
            "company": "Brooker Group",   # stored in description (no native field)
            "subject": "Capital Allocation",
        },
    }
    compose_xlsx(spec, out)

    wb = _load(out)
    assert wb.properties.title == "CAC Monthly Report"
    assert wb.properties.creator == "CAC Agent"
    assert wb.properties.subject == "Capital Allocation"
    # company has no native field — verify it was appended to description
    assert "Brooker Group" in (wb.properties.description or "")


# ---------------------------------------------------------------------------
# 11. Multi-sheet workbook — correct sheet count + independence
# ---------------------------------------------------------------------------


def test_multi_sheet_workbook(tmp_path):
    out = tmp_path / "multi.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Sheet Alpha",
                "headers": ["X"],
                "rows": [["alpha_value"]],
            },
            {
                "name": "Sheet Beta",
                "headers": ["Y"],
                "rows": [["beta_value"]],
            },
            {
                "name": "Sheet Gamma",
                "rows": [["gamma", 42]],
            },
        ]
    }
    result = compose_xlsx(spec, out)
    assert result["sheets"] == 3

    wb = _load(out)
    assert wb.sheetnames == ["Sheet Alpha", "Sheet Beta", "Sheet Gamma"]
    assert wb["Sheet Alpha"]["A2"].value == "alpha_value"
    assert wb["Sheet Beta"]["A2"].value == "beta_value"
    assert wb["Sheet Gamma"]["A1"].value == "gamma"


# ---------------------------------------------------------------------------
# 12. Plain style — no navy fill applied
# ---------------------------------------------------------------------------


def test_plain_style_no_brooker_fill(tmp_path):
    out = tmp_path / "plain.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Plain",
                "headers": ["Col"],
                "rows": [["val"]],
                "style": "plain",
            }
        ]
    }
    compose_xlsx(spec, out)

    wb = _load(out)
    ws = wb["Plain"]

    header_fill = ws["A1"].fill
    # Should NOT have the brooker navy fill
    fg = header_fill.fgColor.rgb.upper() if header_fill.fgColor else "00000000"
    assert fg not in ("FF0F3D5C", "0F3D5C"), (
        "Plain style should not apply navy header fill"
    )


# ---------------------------------------------------------------------------
# 13. Metadata title produces a title row on first sheet only
# ---------------------------------------------------------------------------


def test_metadata_title_on_first_sheet(tmp_path):
    out = tmp_path / "title.xlsx"
    spec = {
        "sheets": [
            {
                "name": "First",
                "headers": ["A", "B"],
                "rows": [["x", "y"]],
            },
            {
                "name": "Second",
                "headers": ["C", "D"],
                "rows": [["p", "q"]],
            },
        ],
        "metadata": {"title": "My Workbook Title"},
    }
    compose_xlsx(spec, out)

    wb = _load(out)
    ws1 = wb["First"]
    ws2 = wb["Second"]

    # Row 1 of first sheet should be the title
    assert ws1["A1"].value == "My Workbook Title"
    # Headers pushed to row 2
    assert ws1["A2"].value == "A"
    assert ws1["B2"].value == "B"
    # Data starts at row 3
    assert ws1["A3"].value == "x"

    # Second sheet should start with headers at row 1 (no title row injected)
    assert ws2["A1"].value == "C"
    assert ws2["A2"].value == "p"


# ---------------------------------------------------------------------------
# 14. Validation — missing required fields
# ---------------------------------------------------------------------------


def test_validate_rejects_missing_sheets():
    with pytest.raises(ValueError, match="sheets"):
        validate_xlsx_spec({})


def test_validate_rejects_empty_sheets():
    with pytest.raises(ValueError, match="sheets"):
        validate_xlsx_spec({"sheets": []})


def test_validate_rejects_sheet_without_name():
    with pytest.raises(ValueError, match="name is required"):
        validate_xlsx_spec({"sheets": [{"rows": [["a"]]}]})


def test_validate_rejects_sheet_without_rows():
    with pytest.raises(ValueError, match="rows is required"):
        validate_xlsx_spec({"sheets": [{"name": "X"}]})


def test_validate_rejects_too_many_sheets():
    spec = {
        "sheets": [{"name": f"S{i}", "rows": [["a"]]} for i in range(25)]
    }
    with pytest.raises(ValueError, match="maximum is 20"):
        validate_xlsx_spec(spec)


# ---------------------------------------------------------------------------
# 15. Sheet name illegal character sanitisation
# ---------------------------------------------------------------------------


def test_sheet_name_illegal_chars_sanitised():
    spec = {
        "sheets": [
            {"name": "Sheet[1]/Test*?", "rows": [["v"]]}
        ]
    }
    normalised = validate_xlsx_spec(spec)
    sanitised = normalised["sheets"][0]["name"]
    # All illegal chars replaced with _
    assert "[" not in sanitised
    assert "]" not in sanitised
    assert "/" not in sanitised
    assert "*" not in sanitised
    assert "?" not in sanitised


# ---------------------------------------------------------------------------
# 16. Column number formats written to data cells
# ---------------------------------------------------------------------------


def test_column_formats_applied(tmp_path):
    out = tmp_path / "formats.xlsx"
    spec = {
        "sheets": [
            {
                "name": "Formatted",
                "headers": ["Label", "Amount", "Pct"],
                "rows": [
                    ["Revenue", 1_250_000.5, 0.1234],
                    ["Costs", 800_000.0, 0.064],
                ],
                "column_formats": {"B": "#,##0.00", "C": "0.0%"},
            }
        ]
    }
    compose_xlsx(spec, out)

    wb = _load(out)
    ws = wb["Formatted"]

    # Data rows start at row 2 (headers at row 1)
    assert ws["B2"].number_format == "#,##0.00"
    assert ws["C2"].number_format == "0.0%"
    assert ws["B3"].number_format == "#,##0.00"
    assert ws["C3"].number_format == "0.0%"


# ---------------------------------------------------------------------------
# 17. Return dict shape
# ---------------------------------------------------------------------------


def test_return_dict_shape(tmp_path):
    out = tmp_path / "shape.xlsx"
    spec = {
        "sheets": [{"name": "S1", "rows": [["a", "b"]]}],
        "charts": [],
    }
    result = compose_xlsx(spec, out)
    assert set(result.keys()) == {"out", "sheets", "cells_written", "charts"}
    assert isinstance(result["out"], str)
    assert isinstance(result["sheets"], int)
    assert isinstance(result["cells_written"], int)
    assert isinstance(result["charts"], int)
