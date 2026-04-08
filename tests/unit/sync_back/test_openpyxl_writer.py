"""Unit tests for sync-back openpyxl_writer module."""
from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pytest
from services.sync_back.src.models import ApprovedProposal
from services.sync_back.src.openpyxl_writer import ExcelWriteError, write_cell

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_proposal(**overrides) -> ApprovedProposal:
    """Return a minimal ApprovedProposal with sensible defaults."""
    defaults = dict(
        proposal_id="chg_0001",
        agent="funding-agent",
        file="ALCO_Tracker.xlsx",
        tab="Sheet1",
        cell="A1",
        old_value="old",
        new_value="new",
        source="Slack #cac-committee",
        confidence=0.95,
        reasoning="test",
        status="approved",
        approved_at=datetime(2026, 4, 1, 12, 0, 0, tzinfo=UTC),
        approved_by="head_of_dept",
    )
    defaults.update(overrides)
    return ApprovedProposal(**defaults)


def _make_excel(
    path: Path,
    sheet_name: str = "Sheet1",
    cell: str = "A1",
    value: object = "old",
) -> None:
    """Create a minimal .xlsx file at *path* with one named sheet and one cell value."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws[cell] = value
    wb.save(str(path))
    wb.close()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestWriteCellUpdatesValue:
    def test_write_cell_updates_value(self, tmp_path: Path) -> None:
        """String new_value is written to the cell and verified on reopen."""
        mirror = tmp_path / "mirror"
        staging = tmp_path / "staging"
        mirror.mkdir()
        staging.mkdir()

        _make_excel(mirror / "ALCO_Tracker.xlsx", sheet_name="Sheet1", cell="A1", value="old")

        proposal = _make_proposal(tab="Sheet1", cell="A1", new_value="new")
        dest = write_cell(proposal, str(mirror), str(staging))

        assert dest.exists()
        wb = openpyxl.load_workbook(str(dest))
        actual = wb["Sheet1"]["A1"].value
        wb.close()
        assert actual == "new"


class TestWriteCellNumericValue:
    def test_write_cell_numeric_value(self, tmp_path: Path) -> None:
        """A new_value of '3.15' is stored as float 3.15."""
        mirror = tmp_path / "mirror"
        staging = tmp_path / "staging"
        mirror.mkdir()
        staging.mkdir()

        _make_excel(
            mirror / "ALCO_Tracker.xlsx",
            sheet_name="Funding Facilities",
            cell="E8",
            value=0.0,
        )

        proposal = _make_proposal(tab="Funding Facilities", cell="E8", new_value="3.15")
        dest = write_cell(proposal, str(mirror), str(staging))

        wb = openpyxl.load_workbook(str(dest))
        actual = wb["Funding Facilities"]["E8"].value
        wb.close()
        assert actual == 3.15
        assert isinstance(actual, float)


class TestWriteCellIntegerValue:
    def test_write_cell_integer_value(self, tmp_path: Path) -> None:
        """A new_value of '42' is stored as int 42."""
        mirror = tmp_path / "mirror"
        staging = tmp_path / "staging"
        mirror.mkdir()
        staging.mkdir()

        _make_excel(mirror / "ALCO_Tracker.xlsx", sheet_name="Sheet1", cell="B2", value=0)

        proposal = _make_proposal(tab="Sheet1", cell="B2", new_value="42")
        dest = write_cell(proposal, str(mirror), str(staging))

        wb = openpyxl.load_workbook(str(dest))
        actual = wb["Sheet1"]["B2"].value
        wb.close()
        assert actual == 42
        assert isinstance(actual, int)


class TestWriteCellNonexistentFileRaises:
    def test_write_cell_nonexistent_file_raises(self, tmp_path: Path) -> None:
        """FileNotFoundError when the source Excel does not exist in mirror."""
        mirror = tmp_path / "mirror"
        staging = tmp_path / "staging"
        mirror.mkdir()
        staging.mkdir()

        # No Excel file created in mirror
        proposal = _make_proposal(file="ALCO_Tracker.xlsx")

        with pytest.raises(FileNotFoundError, match="Source Excel not found"):
            write_cell(proposal, str(mirror), str(staging))


class TestWriteCellWrongTabRaises:
    def test_write_cell_wrong_tab_raises(self, tmp_path: Path) -> None:
        """ExcelWriteError when the requested worksheet does not exist."""
        mirror = tmp_path / "mirror"
        staging = tmp_path / "staging"
        mirror.mkdir()
        staging.mkdir()

        _make_excel(mirror / "ALCO_Tracker.xlsx", sheet_name="Sheet1", cell="A1", value="x")

        proposal = _make_proposal(tab="NonExistentSheet", cell="A1", new_value="y")

        with pytest.raises(ExcelWriteError, match="Worksheet 'NonExistentSheet' not found"):
            write_cell(proposal, str(mirror), str(staging))


class TestWriteCellNeverModifiesMirror:
    def test_write_cell_never_modifies_mirror(self, tmp_path: Path) -> None:
        """The source Excel file in mirror is byte-identical after write_cell."""
        mirror = tmp_path / "mirror"
        staging = tmp_path / "staging"
        mirror.mkdir()
        staging.mkdir()

        source = mirror / "ALCO_Tracker.xlsx"
        _make_excel(source, sheet_name="Sheet1", cell="A1", value="original")

        # Record original bytes
        original_bytes = source.read_bytes()

        proposal = _make_proposal(tab="Sheet1", cell="A1", new_value="modified")
        write_cell(proposal, str(mirror), str(staging))

        # Mirror file must be unchanged
        assert source.read_bytes() == original_bytes

        # Sanity: staging copy has the new value
        dest = staging / "approved" / proposal.proposal_id / "ALCO_Tracker.xlsx"
        wb = openpyxl.load_workbook(str(dest))
        assert wb["Sheet1"]["A1"].value == "modified"
        wb.close()
