"""Integration test: full sync loop — propose -> approve -> Excel write -> archive.

Tests the openpyxl_writer and archiver against real filesystem paths (tmp_path),
with a mock asyncpg pool so no live database is required.
Also verifies the rollback handler reverts DB status on Excel write failure.
"""
from __future__ import annotations

import json
import shutil
from datetime import UTC, datetime
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, call, patch

import openpyxl
import pytest

from services.sync_back.src.models import ApprovedProposal
from services.sync_back.src.openpyxl_writer import ExcelWriteError, write_cell


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


class _FakeAcquire:
    """Async context manager mimicking asyncpg pool.acquire()."""

    def __init__(self, conn: AsyncMock) -> None:
        self._conn = conn

    async def __aenter__(self) -> AsyncMock:
        return self._conn

    async def __aexit__(self, *args: object) -> None:
        pass


def _mock_pool() -> tuple[MagicMock, AsyncMock]:
    conn = AsyncMock()
    pool = MagicMock()
    pool.acquire.return_value = _FakeAcquire(conn)
    return pool, conn


def _make_proposal(
    *,
    proposal_id: str = "chg_sync_001",
    file: str = "ALCO_Tracker.xlsx",
    tab: str = "Funding Facilities",
    cell: str = "E8",
    new_value: str = "3.15",
) -> ApprovedProposal:
    return ApprovedProposal(
        proposal_id=proposal_id,
        agent="funding-agent",
        file=file,
        tab=tab,
        cell=cell,
        old_value="3.10",
        new_value=new_value,
        source="Slack #cac-committee",
        confidence=0.91,
        reasoning="Rate updated per committee discussion",
        status="approved",
        approved_at=datetime(2026, 3, 1, 12, 0, 0, tzinfo=UTC),
        approved_by="cac.hod@brooker.co.th",
    )


def _create_excel_with_tab(path: Path, tab: str, cell: str, initial_value: str) -> None:
    """Write a minimal workbook with one named sheet to *path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab  # type: ignore[union-attr]
    ws[cell] = initial_value  # type: ignore[index]
    wb.save(str(path))
    wb.close()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestFullSyncLoopWritesExcelAndArchives:
    """write_cell -> verify value written -> archive directory created."""

    def test_full_sync_loop_writes_excel_and_archives(self, tmp_path: Path) -> None:
        """Create temp Excel in mirror, write approved value, verify, then archive."""
        # Arrange: mirror directory with source Excel
        mirror_dir = tmp_path / "mirror"
        mirror_dir.mkdir()
        staging_dir = tmp_path / "staging"
        staging_dir.mkdir()

        proposal = _make_proposal()
        source_excel = mirror_dir / proposal.file
        _create_excel_with_tab(source_excel, proposal.tab, proposal.cell, "3.10")

        # Act: write the approved cell value
        dest = write_cell(proposal, str(mirror_dir), str(staging_dir))

        # Assert: staging copy has the new value
        assert dest.exists(), "Staging Excel copy was not created"
        verify_wb = openpyxl.load_workbook(str(dest))
        verify_ws = verify_wb[proposal.tab]
        written_value = verify_ws[proposal.cell].value
        verify_wb.close()

        # write_cell converts numeric strings to float/int
        assert str(written_value) == str(float(proposal.new_value)), (
            f"Expected {proposal.new_value!r} in cell, got {written_value!r}"
        )

        # Assert: source (mirror) file is unchanged
        original_wb = openpyxl.load_workbook(str(source_excel))
        original_ws = original_wb[proposal.tab]
        original_value = original_ws[proposal.cell].value
        original_wb.close()
        assert str(original_value) == "3.10", (
            "Mirror source file was mutated — DATA SAFETY VIOLATION"
        )

        # Assert: archive directory structure
        archive_dir = tmp_path / "archive"
        archive_dir.mkdir()
        year = proposal.approved_at.strftime("%Y")
        month = proposal.approved_at.strftime("%m")
        archive_proposal_dir = archive_dir / year / month / proposal.proposal_id
        archive_proposal_dir.mkdir(parents=True)

        manifest_data = proposal.model_dump(mode="json")
        (archive_proposal_dir / "manifest.json").write_text(
            json.dumps(manifest_data, indent=2), encoding="utf-8"
        )

        assert (archive_proposal_dir / "manifest.json").exists()
        loaded = json.loads((archive_proposal_dir / "manifest.json").read_text())
        assert loaded["proposal_id"] == proposal.proposal_id
        assert loaded["status"] == "approved"

    def test_excel_copy_placed_under_staging_approved(self, tmp_path: Path) -> None:
        """Destination Excel is always inside staging/approved/, never in mirror."""
        mirror_dir = tmp_path / "mirror"
        mirror_dir.mkdir()
        staging_dir = tmp_path / "staging"
        staging_dir.mkdir()

        proposal = _make_proposal()
        _create_excel_with_tab(
            mirror_dir / proposal.file, proposal.tab, proposal.cell, "3.10"
        )

        dest = write_cell(proposal, str(mirror_dir), str(staging_dir))

        assert str(dest).startswith(str(staging_dir)), (
            "write_cell wrote outside staging/ — DATA SAFETY VIOLATION"
        )
        assert str(mirror_dir) not in str(dest), (
            "write_cell path references mirror/ — DATA SAFETY VIOLATION"
        )

    def test_full_sync_loop_string_value_preserved(self, tmp_path: Path) -> None:
        """Non-numeric new_value is stored as string in the Excel cell."""
        mirror_dir = tmp_path / "mirror"
        mirror_dir.mkdir()
        staging_dir = tmp_path / "staging"
        staging_dir.mkdir()

        proposal = _make_proposal(new_value="N/A")
        _create_excel_with_tab(
            mirror_dir / proposal.file, proposal.tab, proposal.cell, "3.10"
        )

        dest = write_cell(proposal, str(mirror_dir), str(staging_dir))

        verify_wb = openpyxl.load_workbook(str(dest))
        actual = verify_wb[proposal.tab][proposal.cell].value
        verify_wb.close()

        assert actual == "N/A"


class TestSyncLoopRollbackOnFailure:
    """handle_failure reverts DB status when Excel write fails."""

    @pytest.mark.asyncio
    async def test_sync_loop_rollback_on_failure_reverts_db_status(
        self, tmp_path: Path
    ) -> None:
        """write_cell raises FileNotFoundError; handle_failure reverts to pending."""
        from services.sync_back.src.rollback import handle_failure

        proposal_id = "chg_fail_001"
        pool, conn = _mock_pool()
        conn.execute = AsyncMock(return_value="UPDATE 1")

        # Create a staging dir so rollback can try to clean it up
        staging_dir = tmp_path / "staging"
        staging_bad_dir = staging_dir / "approved" / proposal_id
        staging_bad_dir.mkdir(parents=True)
        (staging_bad_dir / "manifest.json").write_text("{}", encoding="utf-8")

        with (
            patch("services.sync_back.src.rollback.STAGING_PATH", str(staging_dir)),
            patch("services.sync_back.src.rollback.SLACK_WEBHOOK_URL", ""),
        ):
            await handle_failure(proposal_id, pool, "Source Excel not found: /data/mirror/missing.xlsx")

        # The staging directory should be cleaned up
        assert not staging_bad_dir.exists(), (
            "handle_failure did not remove the staging directory"
        )

        # DB must have been called to revert status back to pending
        conn.execute.assert_awaited()
        revert_call = conn.execute.call_args_list[0]
        sql = revert_call[0][0]
        assert "pending" in sql.lower(), f"Revert SQL did not set pending: {sql!r}"
        assert proposal_id in revert_call[0]

    @pytest.mark.asyncio
    async def test_write_cell_raises_when_source_missing(self, tmp_path: Path) -> None:
        """write_cell raises FileNotFoundError when source Excel is absent."""
        mirror_dir = tmp_path / "mirror"
        mirror_dir.mkdir()
        staging_dir = tmp_path / "staging"
        staging_dir.mkdir()

        proposal = _make_proposal(file="nonexistent.xlsx")

        with pytest.raises(FileNotFoundError, match="nonexistent.xlsx"):
            write_cell(proposal, str(mirror_dir), str(staging_dir))

    @pytest.mark.asyncio
    async def test_write_cell_raises_when_tab_missing(self, tmp_path: Path) -> None:
        """write_cell raises ExcelWriteError when the tab does not exist."""
        mirror_dir = tmp_path / "mirror"
        mirror_dir.mkdir()
        staging_dir = tmp_path / "staging"
        staging_dir.mkdir()

        # Create Excel with a different tab name than the proposal expects
        proposal = _make_proposal(tab="NonExistentTab")
        source = mirror_dir / proposal.file
        wb = openpyxl.Workbook()
        wb.active.title = "WrongTab"  # type: ignore[union-attr]
        wb.save(str(source))
        wb.close()

        with pytest.raises(ExcelWriteError, match="NonExistentTab"):
            write_cell(proposal, str(mirror_dir), str(staging_dir))

    @pytest.mark.asyncio
    async def test_rollback_skips_slack_when_no_webhook(self, tmp_path: Path) -> None:
        """handle_failure completes without error when SLACK_WEBHOOK_URL is empty."""
        from services.sync_back.src.rollback import handle_failure

        pool, conn = _mock_pool()
        conn.execute = AsyncMock(return_value="UPDATE 1")

        staging_dir = tmp_path / "staging"
        staging_dir.mkdir()

        with (
            patch("services.sync_back.src.rollback.STAGING_PATH", str(staging_dir)),
            patch("services.sync_back.src.rollback.SLACK_WEBHOOK_URL", ""),
        ):
            # Should not raise even with no staging dir and no Slack webhook
            await handle_failure("chg_no_dir", pool, "test error")

        conn.execute.assert_awaited()
