"""Excel cell writer using openpyxl — writes approved values to Excel copies."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import structlog

from .models import ApprovedProposal

logger = structlog.get_logger("sync-back.openpyxl_writer")


class ExcelWriteError(Exception):
    """Raised when Excel write or verification fails."""


def write_cell(
    proposal: ApprovedProposal,
    mirror_path: str,
    staging_path: str,
) -> Path:
    """Write proposal's new_value to the correct cell in a COPY of the Excel file.

    Steps:
        1. Locate source Excel: {mirror_path}/{proposal.file}
        2. Copy to {staging_path}/approved/{proposal.proposal_id}/{proposal.file}
        3. Open with openpyxl, navigate to worksheet proposal.tab
        4. Write proposal.new_value to proposal.cell
        5. Save workbook
        6. Verify: reopen, read cell, assert value matches
        7. Return path to modified Excel copy

    NEVER writes to mirror_path (read-only zone).

    Args:
        proposal: The approved proposal with file, tab, cell, new_value.
        mirror_path: Path to /data/mirror (read-only source).
        staging_path: Path to /data/staging (writable).

    Returns:
        Path to the modified Excel file in staging/approved/.

    Raises:
        FileNotFoundError: Source Excel not found in mirror.
        ExcelWriteError: Worksheet not found, write failed, or verification failed.
    """
    # Primary path: direct join (e.g. manifest.file = "alco/ALCO_Tracker.xlsx")
    source = Path(mirror_path) / proposal.file
    if not source.exists():
        # Fallback: search the mirror tree by filename (handles bare filenames
        # like "ALCO_Tracker.xlsx" where the file lives in a subdirectory).
        filename = Path(proposal.file).name
        matches = list(Path(mirror_path).rglob(filename))
        if not matches:
            raise FileNotFoundError(
                f"Source Excel not found in mirror: {source} "
                f"(also searched recursively for '{filename}')"
            )
        source = matches[0]
        logger.info(
            "openpyxl_writer.resolved_via_glob",
            proposal_id=proposal.proposal_id,
            resolved=str(source),
        )

    # Copy to staging/approved/{proposal_id}/
    dest_dir = Path(staging_path) / "approved" / proposal.proposal_id
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / proposal.file
    shutil.copy2(source, dest)

    logger.info(
        "openpyxl_writer.copied",
        source=str(source),
        dest=str(dest),
        proposal_id=proposal.proposal_id,
    )

    # Open and write
    try:
        wb = openpyxl.load_workbook(str(dest))
    except Exception as exc:
        raise ExcelWriteError(f"Failed to open Excel: {exc}") from exc

    if proposal.tab not in wb.sheetnames:
        wb.close()
        raise ExcelWriteError(
            f"Worksheet '{proposal.tab}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[proposal.tab]

    # Write the value — try numeric conversion first
    value: str | float | int = str(proposal.new_value)
    try:
        if "." in str(proposal.new_value):
            value = float(proposal.new_value)
        else:
            value = int(proposal.new_value)
    except (ValueError, TypeError):
        pass  # keep as string

    ws[proposal.cell] = value
    wb.save(str(dest))
    wb.close()

    logger.info(
        "openpyxl_writer.written",
        proposal_id=proposal.proposal_id,
        file=proposal.file,
        tab=proposal.tab,
        cell=proposal.cell,
        value=str(value),
    )

    # Verify: reopen and check
    verify_wb = openpyxl.load_workbook(str(dest))
    verify_ws = verify_wb[proposal.tab]
    actual = verify_ws[proposal.cell].value
    verify_wb.close()

    # Compare as strings for robustness
    if str(actual) != str(value):
        raise ExcelWriteError(
            f"Verification failed for {proposal.cell}: "
            f"expected={value!r}, actual={actual!r}"
        )

    logger.info(
        "openpyxl_writer.verified",
        proposal_id=proposal.proposal_id,
        cell=proposal.cell,
        value=str(value),
    )

    return dest
