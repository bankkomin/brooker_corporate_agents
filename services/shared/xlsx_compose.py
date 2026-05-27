"""Free-form Excel workbook builder.

Given a JSON spec (XlsxComposeSpec), produces a polished .xlsx with multiple
sheets, formulas, formatting, conditional formatting, and optional charts.

This is the *free-form* path — the opposite of the template-fill path in
office_template.py.  Use this when the caller owns the entire structure (column
layout, row data, formulas) and only needs formatting + chart plumbing on top.

Public API
----------
    compose_xlsx(spec, out_path) -> dict
    validate_xlsx_spec(spec)     -> XlsxComposeSpec

Raises ValueError with a message naming the offending sheet/cell on bad input.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Literal, TypedDict

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_BROOKER_NAVY = "0F3D5C"
_BROOKER_NAVY_FULL = "FF0F3D5C"   # openpyxl PatternFill needs ARGB (alpha + RGB)
_BROOKER_EVEN = "FFF2F2F2"
_BROOKER_BORDER = "FFD9D9D9"
_WHITE = "FFFFFFFF"

_MAX_SHEETS = 20
_MAX_CELLS = 100_000
_ILLEGAL_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")

# ---------------------------------------------------------------------------
# TypedDicts (public API contracts)
# ---------------------------------------------------------------------------


class SheetSpec(TypedDict, total=False):
    name: str                          # required; max 31 chars (Excel limit)
    headers: list[str]                 # optional header row
    rows: list[list]                   # required
    formulas: dict[str, str]           # optional {"D2": "=B2*C2"}
    column_widths: dict[str, float]    # optional {"A": 12, "B": 18}
    column_formats: dict[str, str]     # optional {"B": "#,##0.00"}
    freeze_panes: str                  # optional "A2"
    conditional_format: list[dict]     # optional, see _apply_conditional_format
    style: Literal["plain", "brooker"] # default "brooker"


class ChartEmbedSpec(TypedDict, total=False):
    sheet: str                         # target sheet name
    kind: Literal["bar", "line", "pie"]
    data_range: str                    # e.g. "Summary!B2:B10"
    category_range: str                # e.g. "Summary!A2:A10"
    title: str
    anchor: str                        # cell e.g. "F2"


class XlsxComposeSpec(TypedDict, total=False):
    sheets: list[SheetSpec]            # required
    charts: list[ChartEmbedSpec]       # optional
    metadata: dict                     # optional: title, author, company, subject


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def validate_xlsx_spec(spec: dict) -> XlsxComposeSpec:
    """Validate and normalise a raw dict into an XlsxComposeSpec.

    Returns the normalised spec on success.  Raises ValueError with a message
    naming the offending sheet/row on the first violation found.
    """
    if not isinstance(spec, dict):
        raise ValueError("spec must be a dict")

    sheets = spec.get("sheets")
    if not sheets or not isinstance(sheets, list):
        raise ValueError("spec.sheets is required and must be a non-empty list")

    if len(sheets) > _MAX_SHEETS:
        raise ValueError(
            f"spec.sheets has {len(sheets)} entries; maximum is {_MAX_SHEETS}"
        )

    normalised_sheets: list[SheetSpec] = []
    for sheet_idx, raw in enumerate(sheets):
        if not isinstance(raw, dict):
            raise ValueError(f"sheets[{sheet_idx}] must be a dict")

        name: str = raw.get("name", "")  # type: ignore[assignment]
        if not name or not isinstance(name, str):
            raise ValueError(f"sheets[{sheet_idx}].name is required")

        # Sanitise sheet name: replace illegal chars, trim to 31 chars
        sanitised = _ILLEGAL_SHEET_CHARS.sub("_", name)[:31]
        if sanitised != name:
            raw = dict(raw)  # shallow copy so we don't mutate caller's dict
            raw["name"] = sanitised

        rows = raw.get("rows")
        if rows is None or not isinstance(rows, list):
            raise ValueError(f"sheet '{sanitised}': rows is required and must be a list")

        headers = raw.get("headers")
        if headers is not None:
            if not isinstance(headers, list):
                raise ValueError(f"sheet '{sanitised}': headers must be a list")
            for row_idx, row in enumerate(rows):
                if not isinstance(row, list):
                    raise ValueError(
                        f"sheet '{sanitised}' row {row_idx + 1}: expected list, got {type(row).__name__}"
                    )
                if len(row) > len(headers):
                    raise ValueError(
                        f"sheet '{sanitised}' row {row_idx + 1}: "
                        f"{len(row)} cells but headers has {len(headers)} columns"
                    )
        else:
            for row_idx, row in enumerate(rows):
                if not isinstance(row, list):
                    raise ValueError(
                        f"sheet '{sanitised}' row {row_idx + 1}: expected list, got {type(row).__name__}"
                    )

        # Validate formulas dict
        formulas = raw.get("formulas")
        if formulas is not None:
            if not isinstance(formulas, dict):
                raise ValueError(f"sheet '{sanitised}': formulas must be a dict")
            for cell_ref, formula in formulas.items():
                if not isinstance(formula, str) or not formula.startswith("="):
                    raise ValueError(
                        f"sheet '{sanitised}' formula at {cell_ref}: "
                        f"value must be a string starting with '=' (got {formula!r})"
                    )

        # Validate style
        style = raw.get("style", "brooker")
        if style not in ("plain", "brooker"):
            raise ValueError(
                f"sheet '{sanitised}': style must be 'plain' or 'brooker', got {style!r}"
            )

        sheet: SheetSpec = dict(raw)  # type: ignore[assignment]
        sheet["style"] = style  # type: ignore[typeddict-item]
        normalised_sheets.append(sheet)

    # Validate charts
    charts = spec.get("charts", [])
    if not isinstance(charts, list):
        raise ValueError("spec.charts must be a list")

    valid_kinds = ("bar", "line", "pie")
    for ci, chart in enumerate(charts):
        if not isinstance(chart, dict):
            raise ValueError(f"charts[{ci}] must be a dict")
        kind = chart.get("kind")
        if kind not in valid_kinds:
            raise ValueError(
                f"charts[{ci}]: kind must be one of {valid_kinds}, got {kind!r}"
            )
        if not chart.get("sheet"):
            raise ValueError(f"charts[{ci}]: sheet is required")
        if not chart.get("data_range"):
            raise ValueError(f"charts[{ci}]: data_range is required")

    normalised: XlsxComposeSpec = dict(spec)  # type: ignore[assignment]
    normalised["sheets"] = normalised_sheets  # type: ignore[typeddict-item]
    return normalised


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


def _make_header_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type="solid", fgColor=_BROOKER_NAVY_FULL)


def _make_even_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type="solid", fgColor=_BROOKER_EVEN)


def _make_thin_border():
    from openpyxl.styles import Border, Side
    side = Side(style="thin", color=_BROOKER_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_brooker_header(ws, header_row_idx: int, num_cols: int) -> None:
    """Bold white text on navy, centred — applied to the given 1-based row index."""
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = _make_header_fill()
    header_font = Font(name="Calibri", size=11, bold=True, color=_WHITE)
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align


def _apply_brooker_data_rows(
    ws,
    data_start_row: int,
    data_end_row: int,
    num_cols: int,
) -> None:
    """Alternating row bands + thin grey borders on all data cells."""
    even_fill = _make_even_fill()
    border = _make_thin_border()
    base_font = None  # will be set lazily

    from openpyxl.styles import Font

    base_font = Font(name="Calibri", size=11)

    for row_idx in range(data_start_row, data_end_row + 1):
        fill = even_fill if (row_idx % 2 == 0) else None
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.border = border
            cell.font = base_font


def _apply_column_widths(ws, widths: dict[str, float]) -> None:
    from openpyxl.utils import column_index_from_string  # noqa: F401

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter.upper()].width = float(width)


def _apply_column_formats(ws, formats: dict[str, str], data_start_row: int, data_end_row: int) -> None:
    """Apply number format to every data cell in the given column."""
    from openpyxl.utils import column_index_from_string

    for col_letter, fmt in formats.items():
        col_idx = column_index_from_string(col_letter.upper())
        for row_idx in range(data_start_row, data_end_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt


def _apply_conditional_format(ws, rules: list[dict]) -> None:
    """Attach conditional formatting rules to ranges on the worksheet."""
    from openpyxl.formatting.rule import (
        CellIsRule,
        ColorScaleRule,
        DataBarRule,
    )
    from openpyxl.styles import PatternFill

    for rule in rules:
        cell_range: str = rule.get("range", "")
        if not cell_range:
            continue

        rule_type: str = rule.get("type", "")

        if rule_type == "data_bar":
            color = rule.get("color", "638EC6")
            # Normalise common colour names
            _color_map = {
                "blue": "638EC6",
                "green": "63BE7B",
                "red": "F8696B",
                "orange": "FFAC47",
            }
            hex_color = _color_map.get(color.lower(), color).lstrip("#")
            cf_rule = DataBarRule(
                start_type="min",
                start_value=None,
                end_type="max",
                end_value=None,
                color=hex_color,
            )
            ws.conditional_formatting.add(cell_range, cf_rule)

        elif rule_type == "color_scale":
            min_color = rule.get("min_color", "#FFFFFF").lstrip("#")
            max_color = rule.get("max_color", "#0F3D5C").lstrip("#")
            cf_rule = ColorScaleRule(
                start_type="min",
                start_value=None,
                start_color=min_color,
                end_type="max",
                end_value=None,
                end_color=max_color,
            )
            ws.conditional_formatting.add(cell_range, cf_rule)

        elif rule_type == "greater_than":
            value = rule.get("value", 0)
            fill_hex = rule.get("fill", "D9EAD3").lstrip("#")
            fill = PatternFill(fill_type="solid", fgColor=f"FF{fill_hex}")
            cf_rule = CellIsRule(operator="greaterThan", formula=[str(value)], fill=fill)
            ws.conditional_formatting.add(cell_range, cf_rule)

        elif rule_type == "less_than":
            value = rule.get("value", 0)
            fill_hex = rule.get("fill", "F4CCCC").lstrip("#")
            fill = PatternFill(fill_type="solid", fgColor=f"FF{fill_hex}")
            cf_rule = CellIsRule(operator="lessThan", formula=[str(value)], fill=fill)
            ws.conditional_formatting.add(cell_range, cf_rule)

        # Unknown rule types are silently skipped — forward-compat.


# ---------------------------------------------------------------------------
# Sheet builder
# ---------------------------------------------------------------------------


def _write_sheet(
    wb,
    sheet_spec: SheetSpec,
    metadata_title: str | None,
    is_first_sheet: bool,
) -> int:
    """Write one sheet.  Returns number of cells written (data cells only)."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    name: str = sheet_spec["name"]  # type: ignore[typeddict-item]
    headers: list[str] | None = sheet_spec.get("headers")
    rows: list[list] = sheet_spec.get("rows", [])  # type: ignore[assignment]
    formulas: dict[str, str] = sheet_spec.get("formulas", {}) or {}
    column_widths: dict[str, float] = sheet_spec.get("column_widths", {}) or {}
    column_formats: dict[str, str] = sheet_spec.get("column_formats", {}) or {}
    freeze_panes: str | None = sheet_spec.get("freeze_panes")
    cond_fmt: list[dict] = sheet_spec.get("conditional_format", []) or []
    style: str = sheet_spec.get("style", "brooker")  # type: ignore[assignment]

    ws = wb.create_sheet(title=name)

    current_row = 1

    # Title row (only on first sheet when metadata.title is set)
    if is_first_sheet and metadata_title:
        from openpyxl.styles import Alignment, Font

        num_title_cols = max(len(headers) if headers else 0, len(rows[0]) if rows else 1, 1)
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = metadata_title
        title_cell.font = Font(name="Calibri", size=14, bold=True, color=_BROOKER_NAVY)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        if num_title_cols > 1:
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=num_title_cols,
            )
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    header_row_idx: int | None = None
    num_cols = 0

    # Header row
    if headers:
        num_cols = len(headers)
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col_idx).value = h
        header_row_idx = current_row
        current_row += 1

    data_start_row = current_row
    cells_written = 0

    # Data rows
    for row in rows:
        num_cols = max(num_cols, len(row))
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=current_row, column=col_idx).value = value
            cells_written += 1
        current_row += 1

    data_end_row = current_row - 1

    # Formula overrides (written after data so they overwrite any data value)
    for cell_ref, formula in formulas.items():
        ws[cell_ref] = formula
        cells_written += 1

    # Styling
    if style == "brooker" and num_cols > 0:
        if header_row_idx is not None:
            _apply_brooker_header(ws, header_row_idx, num_cols)
        if data_start_row <= data_end_row:
            _apply_brooker_data_rows(ws, data_start_row, data_end_row, num_cols)

    # Column widths
    if column_widths:
        _apply_column_widths(ws, column_widths)

    # Column number formats (data rows only)
    if column_formats and data_start_row <= data_end_row:
        _apply_column_formats(ws, column_formats, data_start_row, data_end_row)

    # Freeze panes
    if freeze_panes:
        ws.freeze_panes = freeze_panes

    # Conditional formatting
    if cond_fmt:
        _apply_conditional_format(ws, cond_fmt)

    return cells_written


# ---------------------------------------------------------------------------
# Chart builder
# ---------------------------------------------------------------------------


def _embed_chart(wb, chart_spec: ChartEmbedSpec) -> None:
    """Add a chart to the named sheet."""
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    target_sheet_name: str = chart_spec["sheet"]  # type: ignore[typeddict-item]
    kind: str = chart_spec.get("kind", "bar")  # type: ignore[assignment]
    data_range: str = chart_spec["data_range"]  # type: ignore[typeddict-item]
    category_range: str | None = chart_spec.get("category_range")
    title: str | None = chart_spec.get("title")
    anchor: str = chart_spec.get("anchor", "F2")  # type: ignore[assignment]

    if target_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"chart specifies sheet '{target_sheet_name}' which does not exist in the workbook"
        )
    ws = wb[target_sheet_name]

    # Build the chart object
    if kind == "bar":
        chart: Any = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
    elif kind == "line":
        chart = LineChart()
    elif kind == "pie":
        chart = PieChart()
    else:
        raise ValueError(f"unsupported chart kind: {kind!r}")

    if title:
        chart.title = title

    # Data reference
    data_ref = Reference(ws, range_string=data_range)
    chart.add_data(data_ref, titles_from_data=True)

    # Category reference
    if category_range:
        cat_ref = Reference(ws, range_string=category_range)
        chart.set_categories(cat_ref)

    chart.shape = 4
    ws.add_chart(chart, anchor)


# ---------------------------------------------------------------------------
# Metadata
# ---------------------------------------------------------------------------


def _apply_metadata(wb, metadata: dict) -> None:
    """Set workbook document properties from the metadata dict.

    openpyxl's DocumentProperties supports: title, creator (author), subject,
    description, keywords, category, lastModifiedBy.  There is no standalone
    'company' field in the core properties — if supplied it is appended to the
    description field so the information is not lost.
    """
    if not metadata:
        return
    props = wb.properties
    if "title" in metadata:
        props.title = str(metadata["title"])
    if "author" in metadata:
        props.creator = str(metadata["author"])
    if "subject" in metadata:
        props.subject = str(metadata["subject"])
    if "description" in metadata:
        props.description = str(metadata["description"])
    if "keywords" in metadata:
        props.keywords = str(metadata["keywords"])
    if "category" in metadata:
        props.category = str(metadata["category"])
    # 'company' is not a standard DocumentProperties field; append to description
    if "company" in metadata:
        company_str = str(metadata["company"])
        existing = props.description or ""
        props.description = f"{existing}; company={company_str}".lstrip("; ")


# ---------------------------------------------------------------------------
# Main entry-point
# ---------------------------------------------------------------------------


def compose_xlsx(spec: XlsxComposeSpec, out_path: str | Path) -> dict:
    """Build a .xlsx file from an XlsxComposeSpec.

    Parameters
    ----------
    spec:
        A dict conforming to XlsxComposeSpec.  Will be validated/normalised
        before use — callers may pass raw dicts from JSON deserialisation.
    out_path:
        Destination file path.  Parent directories are created if needed.

    Returns
    -------
    dict with keys:
        out          — absolute path of the written file (str)
        sheets       — number of sheets created
        cells_written — total data+formula cells written
        charts       — number of charts embedded

    Raises
    ------
    ValueError
        On invalid spec (with a message naming the offending sheet/cell).
    OSError / PermissionError
        On filesystem write failures.
    """
    import openpyxl

    # Validate first — fail fast before touching the filesystem
    normalised = validate_xlsx_spec(dict(spec))

    sheets_spec: list[SheetSpec] = normalised.get("sheets", [])  # type: ignore[assignment]
    charts_spec: list[ChartEmbedSpec] = normalised.get("charts", [])  # type: ignore[assignment]
    metadata: dict = normalised.get("metadata", {}) or {}

    # Cell-count guard (compute before writing)
    total_cell_estimate = sum(
        (len(s.get("headers") or [])) + sum(len(r) for r in (s.get("rows") or []))
        + len(s.get("formulas") or {})
        for s in sheets_spec
    )
    if total_cell_estimate > _MAX_CELLS:
        raise ValueError(
            f"spec would write approximately {total_cell_estimate:,} cells; "
            f"maximum is {_MAX_CELLS:,}"
        )

    wb = openpyxl.Workbook()
    # Remove the default sheet that openpyxl creates
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    total_cells = 0
    metadata_title: str | None = metadata.get("title")

    for sheet_idx, sheet_spec in enumerate(sheets_spec):
        cells = _write_sheet(
            wb=wb,
            sheet_spec=sheet_spec,
            metadata_title=metadata_title,
            is_first_sheet=(sheet_idx == 0),
        )
        total_cells += cells

    # Apply metadata to workbook properties
    _apply_metadata(wb, metadata)

    # Embed charts (after all sheets exist so cross-sheet references resolve)
    charts_embedded = 0
    for chart_spec in charts_spec:
        _embed_chart(wb, chart_spec)
        charts_embedded += 1

    # Write to disk
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    return {
        "out": str(out_path.resolve()),
        "sheets": len(sheets_spec),
        "cells_written": total_cells,
        "charts": charts_embedded,
    }
