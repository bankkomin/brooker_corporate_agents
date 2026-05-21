"""Generate ALCO_Tracker.xlsx from config/excel_schema/alco_tracker.json.

Writes to data/mirror/alco/ALCO_Tracker.xlsx — this is a ONE-TIME bootstrap
script run by a human operator (not an agent).  Agents never call this script;
they only READ from /data/mirror/ and WRITE proposals to /data/staging/pending/.

Usage:
    python scripts/gen_alco_tracker.py
    python scripts/gen_alco_tracker.py --schema config/excel_schema/alco_tracker.json
                                       --out    data/mirror/alco/ALCO_Tracker.xlsx
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark corporate blue
_LABEL_FILL = PatternFill("solid", fgColor="D6E4F0")    # light blue for row labels
_TITLE_FILL = PatternFill("solid", fgColor="2E75B6")    # mid blue for tab title row
_WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_LABEL_FONT = Font(name="Calibri", bold=True, color="1F4E79", size=10)
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_BODY_FONT = Font(name="Calibri", size=10)


def _col_letters(columns: dict[str, str]) -> list[str]:
    """Return sorted column letters for a row spec."""
    return sorted(columns.keys(), key=column_index_from_string)


def _build_sheet(
    wb: openpyxl.Workbook,
    tab: dict,
) -> list[str]:
    """Build one worksheet from a tab spec. Returns list of agent-target cell addresses."""
    ws = wb.create_sheet(title=tab["name"])

    # --- Title row (row 1-6 kept intentionally sparse for Treasury use) ---
    ws.row_dimensions[1].height = 24
    title_cell = ws.cell(row=1, column=2, value=tab["name"])
    title_cell.font = _WHITE_FONT
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- Notes row ---
    ws.cell(row=2, column=2, value=tab.get("notes", "")).font = _BODY_FONT

    # --- Column widths ---
    ws.column_dimensions["A"].width = 4    # gutter
    ws.column_dimensions["B"].width = 22   # Metric / label
    ws.column_dimensions["C"].width = 22   # Threshold / Min / Scenario
    ws.column_dimensions["D"].width = 16   # Current Value
    ws.column_dimensions["E"].width = 16   # Prior / Target

    # --- Header row ---
    header_row = tab["header_row"]  # typically 7

    # Collect all column letters used across all rows for this tab
    all_cols: set[str] = set()
    for row_spec in tab["rows"]:
        all_cols.update(row_spec["columns"].keys())

    for col_letter in sorted(all_cols, key=column_index_from_string):
        # Pick the first row that defines this column's header label
        header_label = ""
        for row_spec in tab["rows"]:
            if col_letter in row_spec["columns"]:
                header_label = row_spec["columns"][col_letter]
                break
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=header_row, column=col_idx, value=header_label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[header_row].height = 18

    # --- Data rows ---
    agent_targets: list[str] = tab.get("agent_targets", [])

    for row_spec in tab["rows"]:
        row_num = row_spec["row"]
        ws.row_dimensions[row_num].height = 16

        for col_letter, _purpose in row_spec["columns"].items():
            col_idx = column_index_from_string(col_letter)
            cell_addr = f"{col_letter}{row_num}"
            cell = ws.cell(row=row_num, column=col_idx)

            if col_letter == "B":
                # Row label
                cell.value = row_spec["label"]
                cell.font = _LABEL_FONT
                cell.fill = _LABEL_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif cell_addr in agent_targets:
                # Agent-writable value cell — leave empty, mark with light border
                cell.value = None
                cell.font = _BODY_FONT
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Thin border to show Treasury where to look
                from openpyxl.styles.borders import Border, Side
                thin = Side(style="thin", color="A0A0A0")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            else:
                # Other data columns — empty placeholder
                cell.value = None
                cell.font = _BODY_FONT
                cell.alignment = Alignment(horizontal="right", vertical="center")

    return agent_targets


def generate(schema_path: Path, out_path: Path) -> None:
    """Build the workbook and write to out_path."""
    with schema_path.open(encoding="utf-8") as fh:
        schema = json.load(fh)

    wb = openpyxl.Workbook()
    # Remove the default blank sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    summary: list[dict] = []

    for tab in schema["tabs"]:
        targets = _build_sheet(wb, tab)
        summary.append(
            {
                "sheet": tab["name"],
                "header_row": tab["header_row"],
                "data_rows": [r["row"] for r in tab["rows"]],
                "agent_target_cells": targets,
            }
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    print(f"Written: {out_path}")
    print(f"Sheets : {[s['sheet'] for s in summary]}")
    print()
    for s in summary:
        print(f"  [{s['sheet']}]")
        print(f"    header_row        : {s['header_row']}")
        print(f"    data_rows         : {s['data_rows']}")
        print(f"    agent_target_cells: {s['agent_target_cells']}")
        print()

    # Verify: reopen and confirm sheet names
    verify_wb = openpyxl.load_workbook(str(out_path))
    assert verify_wb.sheetnames == [t["name"] for t in schema["tabs"]], (
        f"Sheet mismatch: {verify_wb.sheetnames}"
    )
    verify_wb.close()
    print("Verification passed: all sheets present and workbook readable.")


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    default_schema = repo_root / "config" / "excel_schema" / "alco_tracker.json"
    default_out = repo_root / "data" / "mirror" / "alco" / "ALCO_Tracker.xlsx"

    parser = argparse.ArgumentParser(description="Generate ALCO_Tracker.xlsx from schema.")
    parser.add_argument("--schema", type=Path, default=default_schema)
    parser.add_argument("--out", type=Path, default=default_out)
    args = parser.parse_args()

    generate(args.schema, args.out)


if __name__ == "__main__":
    main()
